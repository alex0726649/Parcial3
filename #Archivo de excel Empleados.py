#Archivo de excel Empleados.xlsx
import openpyxl
#Workbook o libro de excel:
wb = openpyxl.load_workbook("Empleados.xlsx")
#WorkSheet u Hoja activa de Excel
ws = wb.active 

def validacion():
    while True:
        try:
            opcion=int(input("\nDigite una opcion \U0001F449: "))
            if 1<=opcion<=7:
                return opcion
            else:
                print("Tienes que seleccionar un número válido")
        except ValueError:
            print("Seleccione una opción")  
    
def ppal(): 
  empleadoData=[]
  opcion=0
  while opcion!=6: 
    print('Bienvenido al programa de administración de empleados') 
    print(" ")
    print("1. Alta")
    print("2. Modificación")
    print("3. Dar de baja un empleado")
    print("4. Consulta de empleado")
    print("5. Sanciones y actualización de sueldo por quincena")
    print("6: Salir del programa")
    print("7: Integrantes del proyecto" )
    opcion=validacion()
    match(opcion):
    #--------------------AGREGAR EMPLEADO----------------------    
        case(1):
            conteof = 0
            #Verificación de la cantidad de datos
            CantMAXfilas=ws.max_row
            CantMAXcol=ws.max_column
            print("\nCantidad de Empleados: ",CantMAXfilas," Número de Datos: ",CantMAXcol)
            
            print("Se agregará un nuevo empleado al registro")
            numeroe=input("\nEscriba el numero del empleado:")
            # -----------VERIFICAR QUE EL EMPLEADO NO EXISTA----------
            columna = ws['A']
            repetido = any(cell.value == numeroe for cell in columna)
            if repetido:
                print('\n\033[91mEste empleado ya existe en la base de datos\033[0m\n\n\n')
                
                
            else:
                nombres=input("Escriba el nombre o nombres (no apellidos) del empleado:")
                apellidos=input("Escriba el o los apellidos del empleado: ")
                print(" se Agrego a : ",numeroe,nombres,apellidos)
                empleadoData.append(numeroe) #NUMERO EMPLEADO
                empleadoData.append(nombres) #NOMBRE
                empleadoData.append(apellidos) #APELLIDO
                empleadoData.append("A")#ESTATUS
                empleadoData.append(1)  #ASISTENCIA 
                empleadoData.append(0)  #RETARDO
                print(empleadoData)
                ws.append(empleadoData)
                wb.save("Empleados.xlsx")  #se guardan los cambios en el archivo
    #----------MODIFICAR ELEMENTO--------------------------------------------------------------------        
        case(2):
            #Modificar un elemento
            CantMAXfilas=ws.max_row
            CantMAXcol=ws.max_column
            conteof=0
            print("\n\033[91mModificarás un elemento\033[0m")
            print("Recuerda que el número de empleado no se podrá modificar")
            numeroe=input("Numero de empleado: ")
            for row in ws:
                num_emplea=row[0].value #primer elemento de una fila
                conteof=conteof+1       #cuanta la cantidad de filas (cantidad de veces que se ejecuta el for)
                if num_emplea==numeroe:
                    print("El empleado se encontró")
                    numfila=conteof
                    #print(numfila)-----------------------------------
                    print("¿Desea modificar...")
                    print("1. Nombre(s)")
                    print("2. Apellidos")
                    print("3. Corregir retraso")
                    print("4. Cambiar Estatus")
                    try:
                        OPmenumodif=int(input("Seleccione una opción a modificar: "))
                        match(OPmenumodif):
                            case(1): 
                                #Modificar nombre
                                nombrenuevo=input("Teclee el(los) nuevo(s) nombre(s): ")
                                ws.cell(row=numfila, column=2, value=nombrenuevo) #accediendo al numero de fila indicado en la columna 2 colocando el valor nuevo en esa celda
                                wb.save("Empleados.xlsx") #se guardan los cambios en el archivo
                                print("El nombre del empleado ",numeroe," se actualizó a ")
                                valornuevo = ws.cell(row=numfila, column=2).value
                                print(valornuevo)
                                print("\U0001F44D")
                                break
                            case(2):
                                #Modificar Apellidos
                                apellidonuevo=input("Teclee el(los) nuevo(s) apellido(s): ")
                                ws.cell(row=numfila, column=3, value=apellidonuevo)
                                wb.save("Empleados.xlsx") #se guardan los cambios en el archivo
                                print("El(los) apellido(s) del empleado ",numeroe," se actualizó a: ")
                                valornuevo = ws.cell(row=numfila, column=3).value
                                print(valornuevo)
                                print("\U0001F44D")
                                break   
                            case(3):
                                #Modificar retardos
                                retardos = ws.cell(row=conteof, column=6).value
                                asistencias = ws.cell(row=conteof, column=5).value
                                print(f'\n\n\033[94mNúmero de retardos: \033[0m{retardos}')
                                if retardos < 1:   
                                    print('Los retardos ya están en el mínimo')

                                else:
                                    print("\033[1mModificación de retardos\033[0m")
                                    ciclo = True
                                    while ciclo:
                                        try:
                                            justificaciones = int(input('\nIngrese el nuevo número de retardos (solamente se pueden disminuir): '))
                                            if justificaciones >= 0 and justificaciones <= 2 and justificaciones != retardos and justificaciones < retardos:
                                                print(f'\n\033[92mEl número de retardos se ha modificado a {justificaciones}\033[0m')
                                                ws.cell(row=conteof, column=6, value=justificaciones)
                                                asistenciaFinal = (retardos - justificaciones) + asistencias
                                                ws.cell(row=conteof, column=5, value=(asistenciaFinal))
                                                wb.save("Empleados.xlsx")
                                                ciclo = False
                                            else:
                                                print('\n\nEl nuevo número de retardos debe tener un rango de 0 a 2 y debe ser menor a el numero de retardos anterior')

                                        except ValueError:
                                            print('Ingrese un valor válido')

                                break
                            case(4):
                                #Modificar Estatus
                                print("Cambiar estatus solo admite A= Activo I=incapacitado")
                                estatus = ws.cell(row=conteof, column=4).value
                                print("El estatus actual del trabajador es: ", estatus)
                                while True:
                                        nestatus = input("Ingresa el nuevo estatus para el empleado A (activo) I (incapacitado): ").upper()
                                        if nestatus in ["A", "I"]:
                                            if nestatus == estatus:
                                                print(f"El empleado ya tiene el estatus '{estatus}'. No se realizó ningún cambio.")
                                            else:
                                                ws.cell(row=conteof, column=4, value=nestatus)
                                                wb.save("Empleados.xlsx")
                                                print("El estatus del trabajador se modificó a: ", nestatus)
                                            break
                                        else:
                                            print("Solo se permite 'A' (Activo) o 'I' (Incapacitado). Intente de nuevo.")
                                 
                            case _:
                                print("\033[91mOpción no válida\033[0m")
                                break
                    except ValueError:
                        print('\n\033[91mIngresa un valor válido\033[0m')
            if conteof==CantMAXfilas:
                    print("\033[94mEmpleado no encontrado \U0001F633\033[0m\n\n")
    #-----------BAJA-----------------------------------------------------------------
        case(3): 
            #BAJA de empleado
            conteof=0
            print("Eliminarás un empleado")
            numeroe=input("Numero de empleado a eliminar \U0001F631: ")
            for row in ws:
                num_emplea=row[0].value #primer elemento de una fila
                conteof=conteof+1       #cuanta la cantidad de filas (cantidad de veces que se ejecuta el for)
                if num_emplea==numeroe:
                    print("El empleado se encontró")
                    numfila=conteof
                   # print(numfila)-----------------------------------------------------
                    ws.cell(row=numfila, column=4, value="B") #"B=BAJA" se asigna un valor en la columna 4
                    wb.save("Empleados.xlsx")#se guardan los cambios en el archivo
                    valornuevo = ws.cell(row=numfila, column=4).value #Se obtiene el valor que se guardo en ESTATUS
                    print("El estatus del empleado ",numeroe," se actualizó a: ",valornuevo)
                    print("\U0001F494")

    #--------------------CONSULTA-----------------------------------------------------------        
        case(4): #Consulta de empleados
            print("\U0001FAF8:")
            CantMAXfilas=ws.max_row
            CantMAXcol=ws.max_column
            conteof=0
            print("Consulta")
            numeroe=input("Numero de empleado a consultar \U0001F9D0: ")
            for row in ws:
                num_emplea=row[0].value #primer elemento de una fila
                conteof=conteof+1       #cuanta la cantidad de filas (cantidad de veces que se ejecuta el for)
                if num_emplea==numeroe:
                        print("El empleado se encontró:")
                        numfila=conteof
                        print('\033[1mRenglón: \033[0m',numfila)
                        break #Ya encontrado el emppleado se interrumpe el for
            informacion = ['','Número de empleado','Nombre(s)','Apellido(s)','Estatus','Asistencias','Retardos','Descuento de sueldo']
            for i in range(1,CantMAXcol): #Desde 1 hasta (1 hasta max cantidad de columnas)
                            valores_empleado=ws.cell(row=numfila,column=i).value 
                            print(f'\033[1m{informacion[i]}:\033[0m {valores_empleado}')
            print('\n\n')
            #Se debe agregar la consulta de todos los empleados (todas las filas y columnas)

#--------------SANCIONES---------------------------------------------------------------------------
        case(5): #Sanciones
            #Se buscarán aquellos que tengan de 3 retardos y se quitará 10% de su sueldo
            print("\U0001FAE2")
            CantMAXfilas=ws.max_row
            CantMAXcol=ws.max_column
            print("Actualizar sanciones")
            i=1
            for i in range(2, CantMAXfilas+1): #Desde 1 hasta (1 hasta max cantidad de columnas)
                cant_retardos=ws.cell(row=i,column=6).value #columna de retardos
                if cant_retardos==3:
                     num_emplea= ws.cell(row=i, column=1).value
                     nombre_emplea= ws.cell(row=i, column=2).value
                     apellidos_emplea= ws.cell(row=i, column=3).value
                     ws.cell(row=i, column=7, value=0.1)#descuento
                     sueldo_actual=21000-(21000*0.1)
                     ws.cell(row=i, column=8, value=sueldo_actual)
                     wb.save("Empleados.xlsx")#se guardan los cambios en el archivo
                     sueldo_nuevo = ws.cell(row=i, column=8).value 
                     print("El empleado ",num_emplea," ",nombre_emplea," ",apellidos_emplea," tiene de sueldo: ",sueldo_nuevo, "\033[1;31mTiene una sanción de 10% porque tiene: ",cant_retardos,'retardos\033[0m')
                else:
                     num_emplea= ws.cell(row=i, column=1).value
                     nombre_emplea= ws.cell(row=i, column=2).value
                     apellidos_emplea=ws.cell(row=i, column=3).value
                     sueldo_actual=21000
                     ws.cell(row=i, column=8, value=sueldo_actual)
                     wb.save("Empleados.xlsx")
                     print("El empleado",num_emplea,"",nombre_emplea,"",apellidos_emplea, "tiene de sueldo: ",sueldo_actual)

        # -------------SALIR DEL PROGRAMA------------------            
        case (6):
            print("Saliendo del programa ")
        case (7):
            print("\n\nEl nombre del equipo es: Equipo Milaneso \U0001F60E")
            print(
                "Los integrantes del equipo son: \n"
                "José Alejandro Pérez Millán-Matricula: 385570\n"
                "Luis Felipe Domínguez Chávez-Matricula: 385500\n"
                "Alexander Chacón Ramírez-Matricula: 385540\n"
                "Diego Flores Verdad Grijalva-Matricula: 385660\n"
            )
            
if __name__=="__main__":
  ppal()  
  #python.exe -m pip install --upgrade pip
  #pip install openpyxl
  #python excelazo.py
  #Si tiene más de 3 retardos, se le quitará un 10% de sueldo
  #Si esta de baja no tendrá ni asistencias ni retardos
  #Si esta incapacitado, tendrá 10 asistencias y 0 retardos

#https://unicode.org/emoji/charts/full-emoji-list.html